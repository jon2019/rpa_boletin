from pathlib import Path
import sys

ROOT_DIR = Path(__file__).resolve().parents[2]
SRC_DIR = ROOT_DIR / 'src'
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

"""Script legacy para exportar diccionario técnico a Excel en docs/legacy/ejecutiva/."""

schema = [
    ('fuentes', [
        ('id', 'SERIAL', '', '', 'PRIMARY KEY', 'YES', 'AUTO_INCREMENT', 'Identificador interno de la fuente'),
        ('url', 'TEXT', '', '', 'UNIQUE', 'YES', '', 'URL del feed RSS o sitio web de la fuente'),
        ('nombre', 'TEXT', '', '', '', 'YES', '', 'Nombre descriptivo de la fuente'),
        ('pais', 'TEXT', '', '', '', 'YES', '', 'Nombre del país de origen'),
        ('metodo', 'VARCHAR', '10', '', '', 'YES', '', 'Método principal de extracción: rss o web scraping'),
        ('scrape_selector', 'TEXT', '', '', '', 'NO', '', 'Selector CSS usado cuando el método es web scraping'),
        ('usuario', 'TEXT', '', '', '', 'NO', '', 'Usuario opcional para login en portales'),
        ('clave', 'TEXT', '', '', '', 'NO', '', 'Clave opcional para login en portales'),
        ('login_url', 'TEXT', '', '', '', 'NO', '', 'URL opcional de login'),
        ('post_login_url', 'TEXT', '', '', '', 'NO', '', 'URL opcional posterior al login'),
        ('activa', 'BOOLEAN', '', '', '', 'YES', 'TRUE', 'Indica si la fuente debe ser procesada'),
    ]),
    ('ejecucion_fuentes', [
        ('id', 'SERIAL', '', '', 'PRIMARY KEY', 'YES', 'AUTO_INCREMENT', 'Identificador interno del resultado de ejecución'),
        ('url_fuente', 'TEXT', '', '', '', 'YES', '', 'URL de la fuente evaluada'),
        ('fecha_ejecucion', 'DATE', '', '', '', 'YES', '', 'Fecha efectiva evaluada en esa corrida'),
        ('nombre_fuente', 'TEXT', '', '', '', 'YES', '', 'Nombre de la fuente cacheado'),
        ('scraping_ok', 'BOOLEAN', '', '', '', 'YES', 'FALSE', 'Indica si scraping/rss se ejecutó correctamente'),
        ('ia_ok', 'BOOLEAN', '', '', '', 'YES', 'FALSE', 'Indica si la etapa de IA se ejecutó correctamente'),
        ('noticias_obtenidas', 'INTEGER', '', '', '', 'YES', '0', 'Cantidad total obtenida de la fuente'),
        ('noticias_enviadas', 'INTEGER', '', '', '', 'YES', '0', 'Cantidad final enviada al resumen'),
        ('error_detalle', 'TEXT', '', '', '', 'NO', '', 'Detalle del error cuando algo falla'),
        ('fecha_registro', 'TIMESTAMP', '', '', '', 'YES', 'CURRENT_TIMESTAMP', 'Fecha de registro técnico'),
    ]),
    ('articulos_pendientes', [
        ('id', 'SERIAL', '', '', 'PRIMARY KEY', 'YES', 'AUTO_INCREMENT', 'Identificador interno del artículo pendiente'),
        ('url_fuente', 'TEXT', '', '', '', 'YES', '', 'Fuente desde la que se obtuvo el artículo'),
        ('fecha_objetivo', 'DATE', '', '', '', 'YES', '', 'Fecha efectiva para la que se guarda'),
        ('titulo', 'TEXT', '', '', '', 'YES', '', 'Título del artículo pendiente'),
        ('url_noticia', 'TEXT', '', '', '', 'YES', '', 'URL del artículo'),
        ('contenido_texto', 'TEXT', '', '', '', 'YES', '', 'Texto preprocesado del artículo'),
    ]),
    ('noticias_enviadas', [
        ('url_hash', 'VARCHAR', '64', '', 'PRIMARY KEY', 'YES', '', 'Hash SHA-256 de la URL para evitar duplicados'),
        ('titulo', 'TEXT', '', '', '', 'YES', '', 'Título de la noticia enviada'),
        ('fuente', 'TEXT', '', '', '', 'YES', '', 'Nombre de la fuente'),
        ('pais', 'TEXT', '', '', '', 'YES', '', 'País de origen'),
        ('url', 'TEXT', '', '', '', 'YES', '', 'URL original de la noticia'),
        ('enviado_en', 'TIMESTAMP', '', '', '', 'YES', 'CURRENT_TIMESTAMP', 'Fecha y hora del envío'),
    ]),
    ('envios_log', [
        ('id', 'SERIAL', '', '', 'PRIMARY KEY', 'YES', 'AUTO_INCREMENT', 'Identificador interno del envío'),
        ('fecha_envio', 'DATE', '', '', '', 'YES', '', 'Fecha efectiva del boletín'),
        ('exito', 'BOOLEAN', '', '', '', 'YES', 'FALSE', 'Resultado del envío SMTP'),
        ('noticias_enviadas', 'INTEGER', '', '', '', 'YES', '0', 'Cantidad enviada en el correo'),
        ('error_detalle', 'TEXT', '', '', '', 'NO', '', 'Detalle del error de envío'),
        ('fecha_registro', 'TIMESTAMP', '', '', '', 'YES', 'CURRENT_TIMESTAMP', 'Fecha de registro técnico'),
    ]),
    ('procesos_programados', [
        ('id', 'SERIAL', '', '', 'PRIMARY KEY', 'YES', 'AUTO_INCREMENT', 'Identificador del proceso programado'),
        ('nombre_proceso', 'TEXT', '', '', 'UNIQUE', 'YES', '', 'Nombre lógico del proceso'),
        ('hora_programada', 'TIME', '', '', '', 'YES', '', 'Hora de ejecución programada'),
        ('activo', 'BOOLEAN', '', '', '', 'YES', 'TRUE', 'Indica si el proceso está activo'),
    ]),
]

OUTPUT_DIR = ROOT_DIR / 'docs' / 'legacy' / 'ejecutiva'
OUTPUT_PATH = OUTPUT_DIR / 'diccionario_boletin.xlsx'


def generar_excel() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Diccionario'

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    section_fill = PatternFill('solid', fgColor='305496')
    section_font = Font(color='FFFFFF', bold=True, size=12)
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['Nombre', 'Tipo', 'Longitud', 'PK/FK', 'Restricción', 'Nullable', 'Default', 'Observación']
    current_row = 1

    for table_name, fields in schema:
        ws.cell(row=current_row, column=1, value=f'Tabla: {table_name}')
        ws.cell(row=current_row, column=1).font = section_font
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        current_row += 1

        for col, value in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border
        current_row += 1

        for field in fields:
            for col, value in enumerate(field, start=1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            current_row += 1

        current_row += 2

    widths = [24, 18, 12, 12, 20, 12, 18, 55]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(OUTPUT_PATH)
    print(f'Excel generado: {OUTPUT_PATH}')
    return OUTPUT_PATH


if __name__ == '__main__':
    generar_excel()
